"""XLSX parser. Walks every sheet; emits one ParsedSegment per data row rendered
as ``header=value | header=value``. Locators carry both ``sheet`` and
``row_index`` so excerpt() can disambiguate rows across sheets.
"""
from pathlib import Path

import openpyxl

from app.schemas import ParsedFile, ParsedSegment


def parse(*, file_id: str, file_name: str, path: Path) -> ParsedFile:
    """Open the workbook read-only, iterate every sheet, and emit one segment per data row."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    segments: list[ParsedSegment] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c is not None else "" for c in rows[0]]
        for r_idx, row in enumerate(rows[1:]):
            cells = [f"{headers[i] if i < len(headers) else ''}={v}" for i, v in enumerate(row)]
            text = " | ".join(cells)
            segments.append(ParsedSegment(
                text=text,
                locator={"type": "xlsx", "sheet": sheet_name, "row_index": r_idx},
            ))
    wb.close()
    return ParsedFile(file_id=file_id, file_name=file_name, type="xlsx", segments=segments)


def excerpt(parsed: ParsedFile, locator: dict) -> str:
    """Return the rendered row matching both ``sheet`` and ``row_index`` in the locator."""
    sheet = locator["sheet"]
    idx = locator["row_index"]
    for seg in parsed.segments:
        if seg.locator["sheet"] == sheet and seg.locator["row_index"] == idx:
            return seg.text
    raise ValueError(f"Sheet={sheet!r} row={idx} not found")
